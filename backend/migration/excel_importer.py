from __future__ import annotations

import hashlib
import json
import re
import uuid
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.models import DailyRecord, MigrationSourceRow
from backend.repositories import EntryRepository
from backend.services.entry_service import DuplicateEntryError, EntryService
from backend.services.schemas import EntryInput, EntryPatch, MedicationInput


SHEET_NAME = "Kopfschmerzkalender"
TABLE_NAME = "KopfschmerzEintraege"
PREVENTIVE_HEADERS = {"Aimovig Spritze", "Momeallerg Nasenspray", "Amitriptylin neuraxpharm"}
NON_CLINICAL_HEADERS = {"Datum", *PREVENTIVE_HEADERS}


@dataclass
class MigrationIssue:
    excel_row: int
    record_date: str | None
    category: str
    message: str


@dataclass
class MigrationReport:
    workbook: str
    started_at: str
    completed_at: str = ""
    workbook_rows: int = 0
    observed_daily_rows: int = 0
    candidate_entries: int = 0
    imported: int = 0
    updated: int = 0
    skipped: int = 0
    duplicates: int = 0
    rejected: int = 0
    daily_records_imported: int = 0
    daily_records_updated: int = 0
    issues: list[MigrationIssue] = field(default_factory=list)
    representative_rows: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["issues"] = [asdict(issue) for issue in self.issues]
        return payload

    def save(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(self.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")


class ExcelImporter:
    def __init__(
        self,
        session: Session,
        user_id: uuid.UUID,
        *,
        annotations: dict[str, dict[str, Any]] | None = None,
        today: date | None = None,
    ) -> None:
        self.session = session
        self.user_id = user_id
        self.repository = EntryRepository(session, user_id)
        self.entry_service = EntryService(session, user_id)
        self.annotations = annotations or {}
        self.today = today or date.today()

    def import_workbook(self, workbook_path: Path) -> MigrationReport:
        workbook_path = workbook_path.resolve()
        report = MigrationReport(workbook=str(workbook_path), started_at=datetime.now(timezone.utc).isoformat())
        rows = self._read_rows(workbook_path)
        report.workbook_rows = len(rows)
        dated_rows = [(excel_row, values) for excel_row, values in rows if isinstance(values.get("Datum"), date)]
        if not dated_rows:
            raise ValueError("Die Excel-Datei enthält keine datierten Kalenderzeilen.")
        cutoff = min(self.today, max(values["Datum"] for _, values in dated_rows))
        rows_in_scope = [(excel_row, values) for excel_row, values in dated_rows if values["Datum"] <= cutoff]
        duplicate_dates = self._duplicate_dates(rows_in_scope)
        handled_dates: set[date] = set()

        for excel_row, values in rows_in_scope:
            record_date = values["Datum"]
            raw_payload = self._json_payload(values)
            content_hash = self._content_hash(raw_payload)
            source_file = workbook_path.name
            source = self.repository.source_row(source_file, SHEET_NAME, excel_row)
            report.observed_daily_rows += 1

            if record_date in duplicate_dates and record_date in handled_dates:
                report.duplicates += 1
                self._record_source(
                    source,
                    source_file,
                    excel_row,
                    record_date,
                    raw_payload,
                    content_hash,
                    "duplicate",
                    ["Datum ist in der Arbeitsmappe mehrfach vorhanden."],
                )
                report.issues.append(MigrationIssue(excel_row, record_date.isoformat(), "duplicate", "Datum ist mehrfach vorhanden."))
                continue
            handled_dates.add(record_date)

            if source is not None and source.content_hash == content_hash and source.status in {"imported", "updated", "daily_only", "skipped"}:
                source.status = "skipped"
                source.imported_at = datetime.now(timezone.utc)
                report.skipped += 1
                continue

            daily_record, daily_created = self._upsert_daily_record(record_date, excel_row, values)
            if daily_created:
                report.daily_records_imported += 1
            else:
                report.daily_records_updated += 1

            entry_id = None
            issues: list[str] = []
            status = "daily_only"
            if self._has_migraine_data(values):
                report.candidate_entries += 1
                try:
                    if not _has_value(values.get("Auslöser (1-6)")):
                        warning = "Auslöser fehlt im Altbestand; als ND (nicht dokumentiert) übernommen."
                        issues.append(warning)
                        report.issues.append(MigrationIssue(excel_row, record_date.isoformat(), "data_quality", warning))
                    payload = self._entry_input(values)
                    annotation = self.annotations.get(record_date.isoformat())
                    existing = self.repository.get_by_date(record_date)
                    if existing is None:
                        entry = self.entry_service.create(payload, origin="excel_migration", reviewed_annotation=annotation)
                        entry.source_row = excel_row
                        entry.source_fingerprint = self._source_fingerprint("entry", record_date)
                        report.imported += 1
                        status = "imported"
                    elif existing.source_system == "excel_migration":
                        entry = self.entry_service.update(
                            existing.id,
                            EntryPatch(**payload.model_dump()),
                            origin="excel_migration",
                        )
                        if annotation:
                            self.entry_service._set_interpretation(entry, reviewed_annotation=annotation)
                        report.updated += 1
                        status = "updated"
                    else:
                        raise DuplicateEntryError("Ein nicht aus Excel stammender Eintrag existiert bereits für dieses Datum.")
                    entry_id = entry.id
                except DuplicateEntryError as exc:
                    report.duplicates += 1
                    status = "duplicate"
                    issues.append(str(exc))
                    report.issues.append(MigrationIssue(excel_row, record_date.isoformat(), "duplicate", str(exc)))
                except (ValueError, InvalidOperation) as exc:
                    report.rejected += 1
                    status = "rejected"
                    issues.append(str(exc))
                    report.issues.append(MigrationIssue(excel_row, record_date.isoformat(), "rejected", str(exc)))

            source = self._record_source(
                source,
                source_file,
                excel_row,
                record_date,
                raw_payload,
                content_hash,
                status,
                issues,
                entry_id=entry_id,
                daily_record_id=daily_record.id,
            )
            if status in {"imported", "updated"} and len(report.representative_rows) < 5:
                report.representative_rows.append(
                    {
                        "excel_row": excel_row,
                        "date": record_date.isoformat(),
                        "strength": values.get("Stärke (0-10)"),
                        "duration_hours": values.get("Dauer (h)"),
                        "trigger": values.get("Auslöser (1-6)"),
                        "notes_length": len(str(values.get("Notizen") or "")),
                        "entry_id": str(entry_id),
                    }
                )

        self.session.flush()
        report.completed_at = datetime.now(timezone.utc).isoformat()
        return report

    @staticmethod
    def _read_rows(workbook_path: Path) -> list[tuple[int, dict[str, Any]]]:
        workbook = load_workbook(workbook_path, read_only=False, data_only=True)
        try:
            sheet = workbook[SHEET_NAME]
            table = sheet.tables[TABLE_NAME]
            min_col, min_row, max_col, max_row = _range_boundaries(table.ref)
            headers = [sheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
            rows = []
            for row_number in range(min_row + 1, max_row + 1):
                values = {str(header): sheet.cell(row_number, column).value for header, column in zip(headers, range(min_col, max_col + 1))}
                if isinstance(values.get("Datum"), datetime):
                    values["Datum"] = values["Datum"].date()
                rows.append((row_number, values))
            return rows
        finally:
            workbook.close()

    @staticmethod
    def _duplicate_dates(rows: list[tuple[int, dict[str, Any]]]) -> set[date]:
        counts: dict[date, int] = {}
        for _, values in rows:
            record_date = values["Datum"]
            counts[record_date] = counts.get(record_date, 0) + 1
        return {record_date for record_date, count in counts.items() if count > 1}

    @staticmethod
    def _has_migraine_data(values: dict[str, Any]) -> bool:
        return any(_has_value(value) for key, value in values.items() if key not in NON_CLINICAL_HEADERS)

    @staticmethod
    def _entry_input(values: dict[str, Any]) -> EntryInput:
        missing = [header for header in ("Stärke (0-10)", "Dauer (h)") if not _has_value(values.get(header))]
        if missing:
            raise ValueError(f"Pflichtfelder fehlen: {', '.join(missing)}")
        medication_name, medication_dose = _split_medication(str(values.get("Medikament") or ""))
        medications = []
        if medication_name:
            medications.append(
                MedicationInput(
                    name=medication_name,
                    dose=medication_dose,
                    effectiveness=_text_or_none(values.get("Hat geholfen?")),
                )
            )
        return EntryInput(
            entry_date=values["Datum"],
            trigger_codes=_split_codes(values.get("Auslöser (1-6)")) or ["ND"],
            strength=int(values["Stärke (0-10)"]),
            duration_hours=Decimal(str(values["Dauer (h)"])),
            pain_type=_text_or_none(values.get("Typ")),
            entered_laterality=_text_or_none(values.get("Seite")),
            aura_codes=_split_codes(values.get("Vorboten (Codes)")),
            vomiting=_marked(values.get("Erbrechen")),
            nausea=_marked(values.get("Übelkeit")),
            phonophobia=_marked(values.get("Lärmscheu")),
            photophobia=_marked(values.get("Lichtscheu")),
            osmophobia=_marked(values.get("Geruchsempfindlich")),
            other_symptom_codes=_split_codes(values.get("Andere Symptome (Codes)")),
            medications=medications,
            notes=str(values.get("Notizen") or ""),
            aimovig_injection=_marked(values.get("Aimovig Spritze")),
            momeallerg_nasal_spray=_marked(values.get("Momeallerg Nasenspray")),
            amitriptyline_neuraxpharm=_marked(values.get("Amitriptylin neuraxpharm")),
        )

    def _upsert_daily_record(self, record_date: date, excel_row: int, values: dict[str, Any]) -> tuple[DailyRecord, bool]:
        record = self.repository.get_daily_record(record_date)
        created = record is None
        if record is None:
            record = DailyRecord(user_id=self.user_id, record_date=record_date)
            self.session.add(record)
        record.aimovig_injection = _marked(values.get("Aimovig Spritze"))
        record.momeallerg_nasal_spray = _marked(values.get("Momeallerg Nasenspray"))
        record.amitriptyline_neuraxpharm = _marked(values.get("Amitriptylin neuraxpharm"))
        record.source_system = "excel_migration"
        record.source_row = excel_row
        record.source_fingerprint = self._source_fingerprint("daily", record_date)
        self.session.flush()
        return record, created

    def _record_source(
        self,
        source: MigrationSourceRow | None,
        source_file: str,
        excel_row: int,
        record_date: date,
        raw_payload: dict[str, Any],
        content_hash: str,
        status: str,
        issues: list[str],
        *,
        entry_id=None,
        daily_record_id=None,
    ) -> MigrationSourceRow:
        if source is None:
            source = MigrationSourceRow(
                user_id=self.user_id,
                source_file=source_file,
                source_sheet=SHEET_NAME,
                source_row=excel_row,
                imported_at=datetime.now(timezone.utc),
                raw_payload={},
                content_hash=content_hash,
                status=status,
            )
            self.session.add(source)
        source.record_date = record_date
        source.raw_payload = raw_payload
        source.content_hash = content_hash
        source.status = status
        source.issues = issues
        source.entry_id = entry_id
        source.daily_record_id = daily_record_id
        source.imported_at = datetime.now(timezone.utc)
        return source

    @staticmethod
    def _json_payload(values: dict[str, Any]) -> dict[str, Any]:
        payload: dict[str, Any] = {}
        for key, value in values.items():
            if isinstance(value, (date, datetime)):
                payload[key] = value.isoformat()
            elif isinstance(value, Decimal):
                payload[key] = float(value)
            else:
                payload[key] = value
        return payload

    @staticmethod
    def _content_hash(payload: dict[str, Any]) -> str:
        canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
        return hashlib.sha256(canonical.encode("utf-8")).hexdigest()

    def _source_fingerprint(self, kind: str, record_date: date) -> str:
        return hashlib.sha256(
            f"excel:{self.user_id}:{kind}:{SHEET_NAME}:{record_date.isoformat()}".encode("utf-8")
        ).hexdigest()


def load_annotations(path: Path | None) -> dict[str, dict[str, Any]]:
    if path is None or not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _range_boundaries(reference: str) -> tuple[int, int, int, int]:
    from openpyxl.utils.cell import range_boundaries

    return range_boundaries(reference)


def _has_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _marked(value: Any) -> bool:
    return str(value or "").strip().lower() in {"x", "ja", "yes", "true", "1"}


def _split_codes(value: Any) -> list[str]:
    return list(dict.fromkeys(part.strip() for part in re.split(r"[,;]", str(value or "")) if part.strip()))


def _text_or_none(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def _split_medication(value: str) -> tuple[str | None, str | None]:
    text = value.strip()
    if not text:
        return None, None
    match = re.fullmatch(r"(.+?)\s*\(([^()]*)\)\s*", text)
    if not match:
        return text, None
    return match.group(1).strip(), match.group(2).strip() or None
